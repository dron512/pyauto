from openpyxl import Workbook
wb = Workbook()
ws = wb.active

scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]

ws.append(('학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사','프로젝트'))

for s in scores:
    ws.append(s)

for idx,score in enumerate(ws['D']):
    if idx == 0:
        continue
    score.value = 10
    # print(idx,score.value)

ws["H1"] = "총점"
ws["I1"] = "성적"

sum = 0
for index,row in enumerate(ws.rows):
    if index == 0 :
        continue
    for cindex,cell in enumerate(row):
        if cindex == 0 :
            continue
        try:
            sum += int(cell.value)
        except:
            continue
    ws["H{}".format(index+1)].value = sum
    if sum >= 90:
        ws["I{}".format(index+1)].value = 'A'
    elif sum >= 80:
        ws["I{}".format(index+1)].value = 'B'
    elif sum >= 70:
        ws["I{}".format(index+1)].value = 'C'
    elif sum >= 60:
        ws["I{}".format(index+1)].value = 'D'
    else:
        ws["I{}".format(index+1)].value = 'F'
    sum = 0

for index,cell in enumerate(ws['B']):
    if index==0 :
        continue
    if cell.value<5:
        print(cell.value)
        ws["I{}".format(index+1)].value = 'F'
wb.save('sample.xlsx')
wb.close()